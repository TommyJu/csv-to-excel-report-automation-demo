from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def format_workbook(wb):

    for ws in wb.worksheets:

        format_headers(ws)

        resize_columns(ws)

        freeze_sheet(ws)



def format_headers(ws):

    header_fill = PatternFill(
        "solid",
        fgColor="DDDDDD"
    )

    border = Border(
        bottom=Side(
            style="thin"
        )
    )


    for row in ws.iter_rows():

        # Detect table headers
        if all(
            cell.value is not None
            for cell in row
        ):

            for cell in row:

                cell.font = Font(
                    bold=True
                )

                cell.alignment = Alignment(
                    horizontal="center"
                )

                cell.fill = header_fill

                cell.border = border



def resize_columns(ws):

    for column in ws.columns:

        max_length = 0

        letter = get_column_letter(
            column[0].column
        )


        for cell in column:

            if cell.value:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )


        ws.column_dimensions[
            letter
        ].width = max_length + 3



def freeze_sheet(ws):

    ws.freeze_panes = "A2"