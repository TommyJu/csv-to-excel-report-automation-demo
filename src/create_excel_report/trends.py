from openpyxl.chart import BarChart, LineChart, Reference


def create_trends_sheet(wb, report):

    ws = wb.create_sheet(
        "Sales Trends"
    )


    # -------------------------
    # Summary Metrics
    # -------------------------

    ws.append([
        "Sales Trend Summary",
        "Value"
    ])

    summary_rows = [
        (
            "Best Sales Date",
            report.time.best_sales_date
        ),

        (
            "Average Daily Sales",
            report.time.average_daily_sales
        )
    ]


    for row in summary_rows:
        ws.append(row)


    ws.append([])


    # -------------------------
    # Monthly Sales
    # -------------------------

    ws.append([
        "Month",
        "Units Sold",
        "Growth %"
    ])


    for month, sales in (
        report.time.sales_by_month.items()
    ):

        growth = (
            report.time.monthly_growth
            .get(month, 0)
        )

        ws.append([
            month,
            sales,
            f"{growth}%"
        ])


    # -------------------------
    # Monthly Sales Chart
    # -------------------------

    chart = LineChart()

    chart.title = (
        "Monthly Sales Trend"
    )

    chart.y_axis.title = (
        "Units Sold"
    )

    chart.x_axis.title = (
        "Month"
    )


    data = Reference(
        ws,
        min_col=2,
        min_row=6,
        max_row=17
    )

    categories = Reference(
        ws,
        min_col=1,
        min_row=6,
        max_row=17
    )


    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(
        categories
    )


    ws.add_chart(
        chart,
        "E2"
    )


    # -------------------------
    # Weekend vs Weekday
    # -------------------------

    start_row = 19


    ws.cell(
        row=start_row,
        column=1,
        value="Day Type"
    )

    ws.cell(
        row=start_row,
        column=2,
        value="Units Sold"
    )


    for index, (day_type, sales) in enumerate(
        report.time.weekend_vs_weekday.items(),
        start=start_row + 1
    ):

        ws.cell(
            row=index,
            column=1,
            value=day_type
        )

        ws.cell(
            row=index,
            column=2,
            value=sales
        )


    chart2 = BarChart()

    chart2.title = (
        "Weekend vs Weekday Sales"
    )


    data = Reference(
        ws,
        min_col=2,
        min_row=start_row,
        max_row=start_row + 2
    )
    
    categories = Reference(
    ws,
    min_col=1,
    min_row=20,
    max_row=21
)


    chart2.add_data(
        data,
        titles_from_data=True
    )
    
    chart2.set_categories(categories)


    ws.add_chart(
        chart2,
        "E20"
    )
    