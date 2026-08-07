from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_DIR = Path("output")
FILE_NAME = "Coffee_Shop_Sales_Report.xlsx"

def create_excel_report(
    report,
    filename=FILE_NAME
):
    
    output_dir = Path("output")

    # Create output folder if missing
    output_dir.mkdir(
        exist_ok=True
    )

    filepath = output_dir / filename
    wb = Workbook()


    create_dashboard(
        wb,
        report
    )

    create_product_sheet(
        wb,
        report
    )

    create_trends_sheet(
        wb,
        report
    )

    create_promotion_sheet(
        wb,
        report
    )


    format_workbook(wb)

    wb.save(filepath)

    print(
        f"Report saved to {filepath}"
    )
    
# Helper functions
def create_dashboard(wb, report):

    ws = wb.active
    ws.title = "Dashboard"


    ws.append([
        "Coffee Shop Sales Dashboard"
    ])

    ws.append([])


    ws.append([
        "Metric",
        "Value"
    ])


    metrics = [
        (
            "Total Items Sold",
            report.summary.total_items_sold
        ),
        (
            "Best Selling Item",
            report.summary.best_selling_item
        ),
        (
            "Best Sales Day",
            report.summary.best_sales_day
        ),
        (
            "Best Sales Month",
            report.summary.best_sales_month
        )
    ]


    for metric, value in metrics:

        ws.append([
            metric,
            value
        ])
        
def create_product_sheet(wb, report):

    ws = wb.create_sheet(
        "Product Performance"
    )


    ws.append([
        "Product",
        "Units Sold"
    ])


    for product, quantity in (
        report.menu.sales_by_menu.items()
    ):

        ws.append([
            product,
            quantity
        ])


    # Add chart

    chart = BarChart()

    chart.title = "Top Products"


    data = Reference(
        ws,
        min_col=2,
        min_row=1,
        max_row=6
    )


    categories = Reference(
        ws,
        min_col=1,
        min_row=2,
        max_row=6
    )


    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(
        categories
    )


    ws.add_chart(
        chart,
        "D2"
    )
    
def create_trends_sheet(wb, report):

    ws = wb.create_sheet(
        "Sales Trends"
    )


    ws.append([
        "Month",
        "Units Sold"
    ])


    for month, sales in (
        report.time.sales_by_month.items()
    ):

        ws.append([
            month,
            sales
        ])


    chart = LineChart()

    chart.title = (
        "Monthly Sales Trend"
    )


    data = Reference(
        ws,
        min_col=2,
        min_row=1,
        max_row=13
    )

    chart.add_data(
        data,
        titles_from_data=True
    )


    ws.add_chart(
        chart,
        "D2"
    )
    
def create_promotion_sheet(wb, report):

    ws = wb.create_sheet(
        "Promotions"
    )


    ws.append([
        "Metric",
        "Value"
    ])


    rows = [
        (
            "Promotion Sales",
            report.promotions.promotion_sales
        ),

        (
            "Regular Sales",
            report.promotions.regular_sales
        ),

        (
            "Promotion Impact %",
            f"{report.promotions.promotion_percentage}%"
        )
    ]


    for row in rows:
        ws.append(row)
        
    
def format_workbook(wb):

    for ws in wb:

        # bold headers

        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )


        # resize columns

        for column in ws.columns:

            max_length = 0

            letter = get_column_letter(
                column[0].column
            )


            for cell in column:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )


            ws.column_dimensions[
                letter
            ].width = max_length + 3